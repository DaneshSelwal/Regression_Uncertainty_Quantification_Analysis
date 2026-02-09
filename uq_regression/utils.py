import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import io
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None

def load_data(train_path, test_path, target_column=None):
    """
    Loads training and testing data from CSV files.

    Args:
        train_path (str): Path to training CSV.
        test_path (str): Path to testing CSV.
        target_column (str, optional): Name of the target column.
                                       If None, the last column is used.
    """
    train_data = pd.read_csv(train_path)
    test_data = pd.read_csv(test_path)

    if target_column:
        y_train = train_data[target_column]
        X_train = train_data.drop(columns=[target_column])
        y_test = test_data[target_column]
        X_test = test_data.drop(columns=[target_column])
    else:
        X_train = train_data.iloc[:, :-1]
        y_train = train_data.iloc[:, -1]
        X_test = test_data.iloc[:, :-1]
        y_test = test_data.iloc[:, -1]

    return X_train, y_train, X_test, y_test

def save_plot_to_excel(fig, excel_file_path, sheet_name):
    """
    Saves a matplotlib figure to an Excel sheet.
    """
    if PILImage is None:
        print("PIL not installed, skipping image save to excel.")
        return

    with io.BytesIO() as buf:
        fig.savefig(buf, format='png')
        buf.seek(0)
        img = OpenpyxlImage(PILImage.open(buf))

        try:
            workbook = load_workbook(excel_file_path)
        except FileNotFoundError:
            # Create a new workbook if it doesn't exist
            from openpyxl import Workbook
            workbook = Workbook()
            # Remove default sheet if creating new
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']

        if sheet_name in workbook.sheetnames:
            # If sheet exists, maybe append specific naming or overwrite?
            # For now let's just use a unique name
            base_name = sheet_name
            counter = 1
            while sheet_name in workbook.sheetnames:
                sheet_name = f"{base_name}_{counter}"
                counter += 1

        worksheet = workbook.create_sheet(title=sheet_name)
        worksheet.add_image(img, 'A1')
        workbook.save(excel_file_path)

def save_values_to_excel(data_dict, excel_file_path, sheet_name):
    """
    Saves a dictionary of values to an Excel sheet.
    """
    df = pd.DataFrame(list(data_dict.items()), columns=['Metric', 'Value'])

    # helper to append to existing excel or create new
    if os.path.exists(excel_file_path):
        with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(excel_file_path, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def plot_metrics_and_save_to_excel(writer, predictions_df, y_test, model_name):
    """
    Placeholder for the plotting function found in notebooks.
    Needs full implementation based on notebook logic if required.
    """
    # This logic was deep in the notebooks, will need to be extracted fully if we want identical behavior.
    # For now, providing a skeleton.
    pass
